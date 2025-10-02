import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from copy import copy
from tqdm import tqdm
import os
import re
import json
import logging

# =============================================================================
# --- CONFIGURACIÓN DEL LOGGING ---
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("registro_proceso.log", mode='w'),
        logging.StreamHandler()
    ]
)

# =============================================================================
# --- FUNCIONES AUXILIARES ---
# =============================================================================
def limpiar_nombre_archivo(nombre):
    return re.sub(r'[\\/*?:"<>|]', "", str(nombre))

def anadir_interactividad_excel(workbook, num_filas, config_interactividad, celda_inicio_str):
    hoja_interactiva = config_interactividad['hoja']
    if hoja_interactiva not in workbook.sheetnames: return
    
    ws = workbook[hoja_interactiva]
    fila_inicio_datos, _ = openpyxl.utils.cell.coordinate_to_tuple(celda_inicio_str)
    fila_encabezados_plantilla = config_interactividad['fila_encabezados']
    headers = [cell.value for cell in ws[fila_encabezados_plantilla]]

    try:
        col_letra_dropdown = get_column_letter(headers.index(config_interactividad['col_dropdown']) + 1)
        col_letra_condicional = get_column_letter(headers.index(config_interactividad['col_condicional']) + 1)
        col_letra_fuente_l1 = get_column_letter(headers.index(config_interactividad['col_fuente_l1']) + 1)
    except ValueError:
        logging.warning(f"No se encontraron las columnas para la interactividad en la hoja '{hoja_interactiva}'.")
        return

    for row in ws.iter_rows():
        for cell in row:
            cell.protection = openpyxl.styles.Protection(locked=False)
    
    dv = DataValidation(type="list", formula1='"L1,L2,L3"', allow_blank=True)
    ws.add_data_validation(dv)
    
    for i in range(fila_inicio_datos, fila_inicio_datos + num_filas):
        celda_dropdown = f"{col_letra_dropdown}{i}"
        celda_condicional = f"{col_letra_condicional}{i}"
        celda_fuente_l1 = f"{col_letra_fuente_l1}{i}"
        formula_excel = f'=IF({celda_dropdown}="L1", {celda_fuente_l1}, IF({celda_dropdown}="L3", 0, ""))'
        ws[celda_condicional].value = formula_excel
        dv.add(ws[celda_dropdown])
        ws[celda_condicional].protection = openpyxl.styles.Protection(locked=True)
    
    ws.protection.sheet = True
    logging.info(f"Interactividad y protección aplicadas a la hoja '{hoja_interactiva}'.")

def copiar_hoja_con_formato(wb_origen, wb_destino, nombre_hoja, estilos_tablas=None):
    try:
        hoja_origen = wb_origen[nombre_hoja]
        hoja_destino = wb_destino.create_sheet(title=nombre_hoja)
    except KeyError:
        logging.warning(f"No se encontró la hoja '{nombre_hoja}' en el origen para copiar.")
        return

    for fila in hoja_origen.iter_rows():
        for celda in fila:
            nueva_celda = hoja_destino.cell(row=celda.row, column=celda.column, value=celda.value)
            if celda.has_style:
                nueva_celda.font, nueva_celda.border, nueva_celda.fill, nueva_celda.number_format, nueva_celda.protection, nueva_celda.alignment = \
                copy(celda.font), copy(celda.border), copy(celda.fill), copy(celda.number_format), copy(celda.protection), copy(celda.alignment)

    for rango in hoja_origen.merged_cells.ranges: hoja_destino.merge_cells(str(rango))
    for col, dim in hoja_origen.column_dimensions.items(): hoja_destino.column_dimensions[col] = copy(dim)
    for row, dim in hoja_origen.row_dimensions.items(): hoja_destino.row_dimensions[row] = copy(dim)

    if estilos_tablas:
        for nombre_tabla, estilo_tabla in estilos_tablas.items():
            if nombre_tabla in hoja_origen.tables:
                tabla_origen = hoja_origen.tables[nombre_tabla]
                nueva_tabla = openpyxl.worksheet.table.Table(displayName=nombre_tabla, ref=tabla_origen.ref)
                nueva_tabla.tableStyleInfo = openpyxl.worksheet.table.TableStyleInfo(name=estilo_tabla, showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                hoja_destino.add_table(nueva_tabla)
            else:
                logging.warning(f"No se encontró la tabla '{nombre_tabla}' en la hoja de origen.")
    logging.info(f"Hoja '{nombre_hoja}' copiada con formato completo.")

# =============================================================================
# --- FUNCIONES DE PROCESAMIENTO MODULARES ---
# =============================================================================

def procesar_un_cost_center(cc, codigo, config):
    logging.info(f"Processing Cost Center: {cc} ({codigo})...")
    
    # Validar archivo Adaptive requerido
    archivo_adaptive = config['archivos']['adaptive_template'].format(codigo=codigo)
    if not os.path.exists(archivo_adaptive):
        logging.warning(f"Omitiendo reporte para '{cc} ({codigo})': No se encontró el archivo requerido '{archivo_adaptive}'.")
        return

    # Preparar archivos de salida y cargar plantilla
    nombre_salida = f"OP2627_{limpiar_nombre_archivo(codigo)}_ScenarioPlanning.xlsx"
    ruta_salida = os.path.join(config['archivos']['salida'], nombre_salida)
    workbook = openpyxl.load_workbook(config['archivos']['plantilla'])

    try: workbook['Overall Planned Cost']['B4'] = codigo
    except KeyError: logging.warning("No se encontró la hoja 'Overall Planned Cost' para escribir el código.")

    # 1. Llenado inicial de datos (Mapeo Principal)
    for hoja_destino, map_config in config['mapeo_principal'].items():
        try:
            df_origen = pd.read_excel(config['archivos']['fuente_principal'], sheet_name=map_config['hoja_origen'], header=map_config['source_header_row'])
            df_filtrado = df_origen[df_origen[config['parametros_globales']['col_agrupacion']] == cc]
            df_para_hoja = df_filtrado[list(map_config['mapeo_columnas'].keys())].copy().rename(columns=map_config['mapeo_columnas'])
            
            hoja_actual = workbook[hoja_destino]
            start_cell = config['plantilla_salida']['celdas_de_inicio'][hoja_destino]
            fila_start, col_start = openpyxl.utils.cell.coordinate_to_tuple(start_cell)

            for r_idx, row_data in enumerate(dataframe_to_rows(df_para_hoja, index=False, header=False), start=fila_start):
                for c_idx, value in enumerate(row_data, start=col_start):
                    hoja_actual.cell(row=r_idx, column=c_idx, value=value)
            
            if config['plantilla_salida']['tabla_bwp'] in hoja_actual.tables and not df_para_hoja.empty:
                tabla = hoja_actual.tables[config['plantilla_salida']['tabla_bwp']]
                nueva_ref = f"{get_column_letter(col_start)}{fila_start-1}:{get_column_letter(col_start + len(df_para_hoja.columns) - 1)}{fila_start + len(df_para_hoja) - 1}"
                tabla.ref = nueva_ref
            
            logging.info(f"Hoja '{hoja_destino}' llenada con {len(df_para_hoja)} filas.")
            if hoja_destino == config['interactividad']['hoja'] and not df_para_hoja.empty:
                anadir_interactividad_excel(workbook, len(df_para_hoja), config['interactividad'], start_cell)
        except Exception as e:
            logging.error(f"Error al procesar el mapeo para '{hoja_destino}'. Error: {e}")

    # 2. Realización de Cálculos Adicionales
    logging.info("Realizando cálculos adicionales...")
    for calculo in config['lista_calculos']:
        try:
            hoja_dest = workbook[calculo['dest_sheet']]
            valor_final = None
            if calculo['op'] == 'SUMIF':
                df_origen = pd.read_excel(config['archivos']['fuente_principal'], sheet_name=calculo['source_sheet'], header=calculo['source_header_row'])
                criterio = cc if calculo['filter_with'] == 'name' else codigo
                valor_final = df_origen.loc[df_origen[calculo['criteria_col']] == criterio, calculo['sum_col']].sum()
            elif calculo['op'] == 'COPY':
                wb_adaptive = openpyxl.load_workbook(archivo_adaptive, data_only=True)
                valor_final = wb_adaptive[calculo['source_sheet']][calculo['source_cell']].value
                wb_adaptive.close()

            if valor_final is not None: hoja_dest[calculo['dest_cell']] = valor_final
        except Exception as e:
            logging.error(f"Error en cálculo para celda {calculo['dest_cell']}. Error: {e}")

    # 3. Copia de la hoja 'Adaptive'
    try:
        wb_origen_adaptive = openpyxl.load_workbook(archivo_adaptive)
        if 'Adaptive' in workbook.sheetnames: del workbook['Adaptive']
        
        estilos_tablas = {tbl: config['plantilla_salida']['estilos_adaptive'] for tbl in config['plantilla_salida']['tablas_adaptive']}
        copiar_hoja_con_formato(wb_origen_adaptive, workbook, 'Adaptive', estilos_tablas)
        wb_origen_adaptive.close()
    except Exception as e:
        logging.error(f"Error al copiar la hoja 'Adaptive'. Error: {e}")
    
    workbook.save(ruta_salida)
    logging.info(f"Reporte guardado en: {ruta_salida}")

# =============================================================================
# --- FUNCIÓN PRINCIPAL ORQUESTADORA ---
# =============================================================================
def main():
    try:
        with open('config.json', 'r', encoding='utf-8') as f:
            config = json.load(f)
    except FileNotFoundError:
        logging.error("No se encontró el archivo 'config.json'. El script no puede continuar.")
        return
        
    if not os.path.exists(config['archivos']['salida']): os.makedirs(config['archivos']['salida'])
    
    logging.info("1. Leyendo hoja maestra para obtener la lista de Cost Centers...")
    try:
        df_maestro = pd.read_excel(config['archivos']['fuente_principal'], sheet_name=config['parametros_globales']['hoja_maestra'], header=config['parametros_globales']['fila_encabezados_maestra'])
        cost_centers = df_maestro[[config['parametros_globales']['col_agrupacion'], config['parametros_globales']['col_codigo']]].dropna().drop_duplicates()
    except Exception as e:
        logging.error(f"No se pudo leer la hoja maestra '{config['parametros_globales']['hoja_maestra']}'. Error: {e}")
        return
    
    logging.info(f"Se encontraron {len(cost_centers)} Cost Centers para procesar.")
    
    for index, fila in tqdm(cost_centers.iterrows(), total=len(cost_centers), desc="Generando Reportes"):
        cc = fila[config['parametros_globales']['col_agrupacion']]
        codigo = fila[config['parametros_globales']['col_codigo']]
        procesar_un_cost_center(cc, codigo, config)

    logging.info("¡Proceso completado!")

if __name__ == "__main__":
    main()